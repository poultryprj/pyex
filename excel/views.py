from django.shortcuts import render
import json
from openpyxl import load_workbook
from django.http import JsonResponse
from rest_framework.decorators import api_view
from openpyxl.utils.dataframe import dataframe_to_rows
from rest_framework import status
import pandas as pd
from openpyxl import load_workbook, Workbook
from rest_framework.response import Response
from .models import ExcelData  # Import your ExcelData model
from datetime import datetime  # Import datetime module
import pytz  # Import pytz module
from openpyxl.styles import Alignment
from datetime import datetime, timedelta
import openpyxl
from openpyxl.worksheet.views import SheetView, Selection
import os
from openpyxl.utils import get_column_letter


@api_view(['POST'])
def excel_view(request, sheet_name):
    if request.method == 'POST':
        try:
            data = json.loads(request.body.decode("utf-8"))
            json_objects = data.get('json_objects')

            # Provide the full path to your Excel file
            file_path = 'main.xlsx'

            if not json_objects:
                return Response({"error": "JSON objects are required."}, status=status.HTTP_400_BAD_REQUEST)

            # Define the default columns outside of the if-else block
            default_columns = ['   date   ', '   time   ', 'shop_code', 'product_type',
                               'product_id',  'weight', 'quantity', 'daily_rate', 'rate', 'amount']

            # Load the Excel workbook using openpyxl
            workbook = load_workbook(filename=file_path)

            # Find the current sheet or create it if it doesn't exist
            if sheet_name not in workbook.sheetnames:
                if len(workbook.sheetnames) == 0:
                    # Create the first sheet
                    new_sheet_name = "Raw_data_01"
                    new_sheet = workbook.create_sheet(title=new_sheet_name)
                    new_sheet.append(['Raw Data Summary'])  # Add the title row
                    new_sheet.append(default_columns)  # Add the column names

                    # Set column widths for default columns
                    for column_letter, column_name in zip('ABCDEFGHIJKLM', default_columns):
                        column = new_sheet.column_dimensions[column_letter]
                        # Adjust the width as needed
                        column.width = len(column_name) + 2

                    # Set the alignment for the header row (centered)
                    header_row = new_sheet[2]
                    for cell in header_row:
                        cell.alignment = Alignment(horizontal='center')
                else:
                    # Find the next available sheet name with a sequential number
                    sheet_number = 1
                    while True:
                        new_sheet_name = f"Raw_data_{sheet_number:02d}"
                        if new_sheet_name not in workbook.sheetnames:
                            break
                        sheet_number += 1
                    new_sheet = workbook.create_sheet(title=new_sheet_name)
                    new_sheet.merge_cells('A1:J1')
                    title_cell = new_sheet.cell(row=1, column=1)
                    title_cell.value = 'RAW DATA'
                    title_cell.alignment = Alignment(horizontal='center')
                    new_sheet.append(default_columns)  # Add the column names

                    # Set column widths for default columns
                    for column_letter, column_name in zip('ABCDEFGHIJKLM', default_columns):
                        column = new_sheet.column_dimensions[column_letter]
                        # Adjust the width as needed
                        column.width = len(column_name) + 2

                    # Set the alignment for the header row (centered)
                    header_row = new_sheet[2]
                    for cell in header_row:
                        cell.alignment = Alignment(horizontal='center')

                sheet = new_sheet
            else:
                sheet = workbook[sheet_name]

            # Check if column names already exist in the sheet
            column_names = [cell.value for cell in sheet[2]]

            # Get the current date and time in UTC+05:30 (IST)
            # Define the IST timezone
            ist = pytz.timezone('Asia/Kolkata')

            # Get the current date and time in the IST timezone
            current_datetime = datetime.now(ist)

            # Format current_date as "dd/mm/yyyy" and current_time as "HH:mm:ss"
            current_date = current_datetime.strftime(
                '%d-%m-%Y')  # Format date as dd/mm/yyyy
            current_time = current_datetime.strftime(
                '%H:%M:%S')  # Format time as HH:mm:ss

            # Append data only if the column names match and product_id is match 1 to 25
            if column_names == default_columns:
                # ...
                # Inside the loop that processes JSON objects
                for obj in json_objects:
                    product_id = obj.get('product_id', 0)
                    if product_id not in range(1, 29):
                        return Response({"error": "Please enter a valid product_type (>=1 To <=28)"}, status=status.HTTP_400_BAD_REQUEST)

                    if product_id in [4, 5, 12, 19, 22, 23, 24, 25, 26, 27, 28]:
                        # product_id == 4 or 5 weight not needed
                        weight = obj.get('weight')
                        if weight:
                            return Response({"error": "please remove weight field..!!"}, status=status.HTTP_400_BAD_REQUEST)
                    else:
                        # weight needed
                        weight = obj.get('weight')
                        if not weight:
                            return Response({"error": "please enter weight..!!"}, status=status.HTTP_400_BAD_REQUEST)

                    # Handle the weight when appending to the sheet
                    if weight == "":
                        row = [current_date, current_time, obj.get('shop_code', 0), obj.get('product_type', 0), obj.get('product_id', 0), "", obj.get(
                            'quantity', 0.0), obj.get('daily_rate', 0.0), obj.get('rate', 0.0), ""]
                    else:
                        row = [current_date, current_time, obj.get('shop_code', 0), obj.get('product_type', 0), obj.get('product_id', 0), weight, obj.get(
                            'quantity', 0.0), obj.get('daily_rate', 0.0), obj.get('rate', 0.0), ""]

                    # Append data to the sheet
                    sheet.append(row)
                    sheet.freeze_panes = "A3"

                    # Create an Alignment object to center align text
                    alignment = Alignment(horizontal='center')

                    # Apply the alignment to all cells in the last row of the sheet
                    for cell in sheet[sheet.max_row]:
                        cell.alignment = alignment

                # Save the updated Excel file after appending data
                workbook.save(file_path)

                # the formula
                formula = '=IF(OR(E{i}=1, E{i}=2, E{i}=3, E{i}=8, E{i}=6, E{i}=7, E{i}=9, E{i}=10, E{i}=14, E{i}=13, E{i}=11, E{i}=15, E{i}=16, E{i}=17, E{i}=18, E{i}=20, E{i}=21), F{i}*H{i}, IF(OR(E{i}=4, E{i}=5, E{i}=12, E{i}=19, E{i}=22, E{i}=23, E{i}=24, E{i}=25, E{i}=26, E{i}=27, E{i}=28), G{i}*H{i}, ""))'

                # Loop through the rows and apply the dynamic formula
                for i in range(3, sheet.max_row + 1):
                    # Replace {i} with the current row number
                    row_formula = formula.format(i=i)
                    sheet.cell(row=i, column=default_columns.index(
                        "amount") + 1).value = row_formula

                # Save the Excel file again after applying the formula
                workbook.save(file_path)

                return Response({'message': 'Data appended successfully'}, status=status.HTTP_200_OK)
            else:
                return Response({"error": "Column names in the data do not match the existing sheet"}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


################################


@api_view(['POST'])
def create_daily_summary_sheet(request, sheet_name):
    if request.method == 'POST':
        try:
            data = json.loads(request.body.decode("utf-8"))

            # Provide the full path to your Excel file
            file_path = 'main.xlsx'

            # Check if the file exists
            if not os.path.isfile(file_path):
                return JsonResponse({'error': f'File not found at path: {file_path}'}, status=status.HTTP_400_BAD_REQUEST)

            # Load the Excel workbook using openpyxl
            workbook = load_workbook(filename=file_path)

            if sheet_name in workbook.sheetnames:
                return JsonResponse({'error': f'Sheet "{sheet_name}" already exists'}, status=status.HTTP_400_BAD_REQUEST)

            # Create a new sheet with the provided sheet_name
            new_sheet = workbook.create_sheet(title=sheet_name)

            # # Define the default columns and add them to the A2 row
            # Define the product types and their respective column ranges
            product_types = {
                'ACCOUNT': (1, 4),
                'LARGE BOILER': (6, 11),
                'SMALL BOILER': (13, 18),
                'GAVRAN': (20, 25),
                'DUCK': (27, 32),
                'KADAKNATH': (34, 39),
                'BATER' : (41,46),
                'WHITE EGGS' : (48,52),
                'BROWN EGGS' : (54,58),
                'SURMAY' : (60,65),
                'PAPLETE':(67,72),
                'BUMLA' : (74,79),    #3,14 WT
                'PRAWNS' : (81,86),
                'BANGDA' : (88,93),
                'ROHU' : (95,100),
                'CRAB' : (102,106),   #3,12
                'MENDNI' : (108,113),
                'SHELI' : (115,120),
                'PACKED CHICKEN' : (122,127),
                'PACKED EGGS' : (129,133),   #5,19
                'PACKED FISH' : (135,140),
                'PACKED MUTTON' : (142,147),
                'DOG FOOD' : (149,153),  #6,22
                'CAT FOOD' : (155,159),
                'FISH FOOD' : (161,165),
                'MUTTON MASALA' : (167,171),
                'EGGS MASALA' : (173,177),
                'FISH MASALA' : (179,183),
                'KOLHAPURI MASALA' : (185,189),

                # Add more product types as needed
            }

            default_columns = [
                'Date',
                'opening_balance',
                'paid_amount',
                'closing_balance',
                ''
            ]

            total_products = [1, 2, 3, 8, 6, 7, 4, 5, 9, 10, 14, 13, 11,
                              15, 12, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27, 28]
            list_1 = [4, 5, 12, 19, 22, 23, 24, 25, 26, 27, 28]

            for product_id in total_products:
                if product_id not in list_1:
                    default_columns.extend([
                        'product_type',
                        'product_id',
                        'weight',
                        'quantity',
                        'rate',
                        'amount',
                        ''
                    ])
                if product_id in list_1:
                    default_columns.extend([
                        'product_type',
                        'product_id',
                        'quantity',
                        'rate',
                        'amount',
                        ''
                    ])

            # Merge cells and set titles for each product type dynamically
            for product_type, (start_col, end_col) in product_types.items():
                new_sheet.merge_cells(
                    start_row=1, start_column=start_col, end_row=1, end_column=end_col)
                title_cell = new_sheet.cell(row=1, column=start_col)
                title_cell.value = product_type
                title_cell.alignment = Alignment(horizontal='center')

            for col_num, header in enumerate(default_columns, start=1):
                new_sheet.cell(row=2, column=col_num, value=header)

            # Set the alignment for the header row (centered)
            header_row = new_sheet[2]
            for cell in header_row:
                cell.alignment = Alignment(horizontal='center')

            # Set column widths for default columns
            for i, column_name in enumerate(default_columns):
                # +1 because columns are 1-indexed
                column_letter = get_column_letter(i + 1)
                column = new_sheet.column_dimensions[column_letter]
                # Adjust the width as needed
                # Minimum width of 12
                column.width = max(len(column_name) + 2, 12)

            # Define the financial year start and end dates
            financial_year_start = datetime(2023, 4, 1)
            financial_year_end = datetime(2024, 3, 31)

            # Iterate over each date within the financial year
            current_date = financial_year_start
            while current_date <= financial_year_end:
                # Create a new row for each date
                row = [current_date.strftime('%d-%m-%Y'), '', '', '', '', 1, 1, '', '', '', '', '', 1, 2, '', '', '', '', '', 1, 3, '', '', '', '', '', 1, 8,
                       '', '', '', '', '', 1, 6, '', '', '', '', '', 1, 7, '', '', '', '', '', 2, 4, '', '', '', '', 2, 5, '', '', '', '',
                       3, 9, '', '', '', '', '', 3, 10, '', '', '', '', '', 3, 14, '', '', '', '', '', 3, 13, '', '', '', '', '', 3, 11, '', '', '', '', '', 3, 15, '', '', '', '', '', 3, 12,
                       '', '', '', '', 4, 16, '', '', '', '', '', 4, 17, '', '', '', '', '', 5, 18, '', '', '', '', '', 5, 19, '', '', '', '', 5, 20, '', '', '', '', '', 5, 21, '', '', '', '', '',
                       6, 22, '', '', '', '', 6, 23, '', '', '', '', 6, 24, '', '', '', '', 7, 25, '', '', '', '', 7, 26, '', '', '', '', 7, 27, '', '', '', '', 7, 28, '', '', '']
                new_sheet.append(row)

                # Move to the next date
                current_date += timedelta(days=1)
            workbook.save(file_path)


            weight_products = { 
                # For weight products
                1: ('H', 'I', 'J', 'K'),     # 1 of 1
                2: ('O', 'P', 'Q', 'R'),     # 1 of 2
                3: ('V', 'W', 'X', 'Y'),     # 1 of 3
                8: ('AC', 'AD', 'AE', 'AF'), # 1 of 8
                6: ('AJ', 'AK', 'AL', 'AM'), # 1 of 6
                7: ('AQ', 'AR', 'AS', 'AT'), # 1 of 7
                9: ('BJ', 'BK', 'BL', 'BM'), # 3 of 9 surmay
                10: ('BQ', 'BR', 'BS', 'BT'), # 3 of 10 paplet
                14: ('BX', 'BY', 'BZ', 'CA'), # 3 of 14 bumla
                13: ('CE', 'CF', 'CG', 'CH'), # 3 of 13 prawns
                11: ('CL', 'CM', 'CN', 'CO'), # 3 of 11 bangda
                15: ('CS', 'CT', 'CU', 'CV'), # 3 of 15 rohu
                16: ('DF', 'DG', 'DH', 'DI'), # 4 of 16 mendni
                17: ('DM', 'DN', 'DO', 'DP'), # 4 of 17 sheli
                18: ('DT', 'DU', 'DV', 'DW'), # 5 of 18 pkd chicken
                20: ('EG', 'EH', 'EI', 'EJ'), # 5 of 20 pkd fish
                21: ('EN', 'EO', 'EP', 'EQ'), # 5 of 21 pkd mutton


            }

            quantity_products = {
                # For Quantity Products
                4: ('AX', 'AY', 'AZ'), # 2 of 4 WHITE EGGS
                5: ('BD', 'BE', 'BF'), # 2 of 5 BROWN EGGS
                12: ('CZ', 'DA', 'DB'), # 3 of 12 CRAB
                19: ('EA', 'EB', 'EC'), # 5 of 19 PKD EGGS
                22: ('EU', 'EV', 'EW'), # 6 of 22 DOG FOOD
                23: ('FA', 'FB', 'FC'), # 6 of 23 CAT FOOD
                24: ('FG', 'FH', 'FI'), # 6 of 24 FISH FOOD
                25: ('FM', 'FN', 'FO'), # 7 of 25 MUTTON MASALA
                26: ('FS', 'FT', 'FU'), # 7 of 26 EGGS MASALA
                27: ('FY', 'FZ', 'GA'), # 7 of 27 FISH MASALA
                28: ('GE', 'GF', 'GG'), # 7 of 28 KOLHAPURI MASALA
            }

            for product_id, (weight_col, quantity_col, rate_col, amount_col) in weight_products.items():
                for row in range(3, 369): 
                        # For Weight Products
                        avg_weight = f'=IF(COUNTIFS(Raw_data_01!A:A,$A{row},Raw_data_01!E:E,{product_id})>0,SUMIFS(Raw_data_01!F:F,Raw_data_01!A:A,$A{row},Raw_data_01!E:E,{product_id}), "")'
                        new_sheet[f'{weight_col}{row}'] = avg_weight

                        sum_of_quantity = f'=IF(COUNTIFS(Raw_data_01!A:A,$A{row},Raw_data_01!E:E,{product_id})>0,SUMIFS(Raw_data_01!G:G,Raw_data_01!A:A,$A{row},Raw_data_01!E:E,{product_id}), "")'
                        new_sheet[f'{quantity_col}{row}'] = sum_of_quantity

                        avg_rate = f'=IF(COUNTIFS(Raw_data_01!A:A,$A{row},Raw_data_01!E:E,{product_id})>0,AVERAGEIFS(Raw_data_01!I:I,Raw_data_01!A:A,$A{row},Raw_data_01!E:E,{product_id}), "")'
                        new_sheet[f'{rate_col}{row}'] = avg_rate

                        sum_of_amount = f'=IF(COUNTIFS(Raw_data_01!A:A,$A{row},Raw_data_01!E:E,{product_id})>0,SUMIFS(Raw_data_01!J:J,Raw_data_01!A:A,$A{row},Raw_data_01!E:E,{product_id}), "")'
                        new_sheet[f'{amount_col}{row}'] = sum_of_amount

            for product_id, (quantity_col, rate_col, amount_col) in quantity_products.items():
                for row in range(3, 369):
                    # For Quantity Products
                    sum_of_quantity = f'=IF(COUNTIFS(Raw_data_01!A:A,$A{row},Raw_data_01!E:E,{product_id})>0,SUMIFS(Raw_data_01!G:G,Raw_data_01!A:A,$A{row},Raw_data_01!E:E,{product_id}),"")'
                    new_sheet[f'{quantity_col}{row}'] = sum_of_quantity

                    avg_rate = f'=IF(COUNTIFS(Raw_data_01!A:A,$A{row},Raw_data_01!E:E,{product_id})>0,AVERAGEIFS(Raw_data_01!I:I,Raw_data_01!A:A,$A{row},Raw_data_01!E:E,{product_id}),"")'
                    new_sheet[f'{rate_col}{row}'] = avg_rate

                    sum_of_amount = f'=IF(COUNTIFS(Raw_data_01!A:A,$A{row},Raw_data_01!E:E,{product_id})>0,SUMIFS(Raw_data_01!J:J,Raw_data_01!A:A,$A{row},Raw_data_01!E:E,{product_id}),"")'
                    new_sheet[f'{amount_col}{row}'] = sum_of_amount

            # cloasing balance
            for row in range(3, 369):
                closing_balance = f'=SUM(B{row},K{row},R{row},Y{row},AF{row},AM{row},AT{row},BM{row},BT{row},CA{row},CH{row},CO{row},CV{row},DI{row},DP{row},DW{row},EJ{row},EQ{row},AZ{row},BF{row},DB{row},EC{row},EW{row},FC{row},FI{row},FO{row},FU{row},GA{row},GG{row}) - C{row}'
                new_sheet[f'D{row}'] = closing_balance

            # Add the formula to the "B" column (opening_balance column) from B4 to B368
            for row in range(4, 369):
                formula = f'=IF(D{row - 1}<>0, D{row - 1}, IFERROR(INDEX(D3:D${row - 1}, MATCH(1, D3:D${row - 1}<>0, 0)), LOOKUP(2, 1/(D3:D${row - 1}<>0), D3:D${row - 1})))'
                new_sheet[f'B{row}'] = formula
            # Define a function to convert Excel column letters to column index
            def col_letter_to_index(col_letter):
                result = 0
                for letter in col_letter:
                    result = result * 26 + (ord(letter) - ord('A') + 1)
                return result

            # Format the columns
            columns_to_format = ['B', 'C', 'D', 'H', 'J', 'K', 'O', 'Q', 'R', 'V', 'X', 'Y', 'AC', 'AE', 'AF', 'AJ', 'AL', 'AM', 'AQ', 'AS', 'AT', 'AY', 'AZ', 'BE', 'BF', 'BJ', 'BL', 'BM', 'BQ', 'BS', 'BT', 'BX', 'BZ', 'CA', 'CE', 'CG', 'CH', 'CL', 'CN', 'CO', 'CS', 'CU', 'CV', 'DA', 'DV', 'DF', 'DH', 'DI', 'DM', 'DO', 'DP', 'DT', 'DV', 'DW', 'EB', 'EC', 'EG', 'EI', 'EJ', 'EN', 'EP', 'EQ', 'EV', 'EW', 'FB', 'FC', 'FH', 'FI', 'FN', 'FO', 'FT', 'FU', 'FZ', 'GA', 'GF', 'GG']

            for col_letter in columns_to_format:
                col_index = col_letter_to_index(col_letter)
                # Format the columns to display two decimal places
                for row in new_sheet.iter_rows(min_row=3, max_row=369, min_col=col_index, max_col=col_index):
                    for cell in row:
                        cell.number_format = '0.00'

            # Freeze the top row (column names) when scrolling
            new_sheet.freeze_panes = "A3"

            # Save the updated Excel file again
            workbook.save(file_path)

            return JsonResponse({'message': f'Successfully Created {sheet_name} with Data & Formulas'}, status=status.HTTP_200_OK)
        except FileNotFoundError as e:
            return JsonResponse({'error': 'File not found'}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    else:
        return JsonResponse({'error': 'Invalid request method'}, status=status.HTTP_400_BAD_REQUEST)
